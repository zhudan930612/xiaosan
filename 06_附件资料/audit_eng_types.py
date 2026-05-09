#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
基于条款级底稿，全面核对导入模板中每个组织的可审核工程类型
找出所有遗漏和错误
"""

import openpyxl
import os

# 条款到工程类型名称的映射
CLAUSE_TO_TYPE = {
    "T3-1": "房屋建筑",
    "T3-2": "交通工程",
    "T3-3": "水务工程",
    "T3-4": "园林绿化工程",
    "T3-5": "既有建筑修缮",
    "T3-6": "建筑装饰装修工程",
    "T3-7": "既有建筑加装电梯工程",
    "T3-8": "临时建设",
    "T3-9": "农民自建房",
    "T3-10": "农业设施建设工程",
    "T3-11": "农村地区综合性建设工程",
    "T3-12": "人防工程",
    "T4": "住宅室内装饰装修工程",
}

# 从条款级底稿提取的每个组织期望的可审核工程类型（新全称 → 类型列表）
# 注：T3-5/T3-7的住宅/非住宅细分在底稿中体现，但导入模板中不区分，统一用条款名
EXPECTED_TYPES = {
    # 西虹桥
    "西虹桥公司-建设管理部": ["房屋建筑", "交通工程", "人防工程", "园林绿化工程", "建筑装饰装修工程", "既有建筑修缮", "既有建筑加装电梯工程", "临时建设"],

    # 华新镇
    "华新镇-城市建设管理事务中心": ["住宅室内装饰装修工程", "既有建筑修缮", "既有建筑加装电梯工程", "水务工程"],
    "华新镇-规划建设和生态环境办公室": ["房屋建筑", "交通工程", "人防工程", "农民自建房", "建筑装饰装修工程", "既有建筑修缮", "临时建设"],
    "区市场监管局-华新市场监督管理所": ["既有建筑加装电梯工程"],
    "华新镇-镇属企业-绿化公司": ["园林绿化工程"],
    "华新镇-农业农村服务中心": ["农业设施建设工程", "农村地区综合性建设工程"],

    # 朱家角镇
    "朱家角镇-城市建设管理事务中心": ["住宅室内装饰装修工程", "水务工程"],
    "朱家角镇-规划建设和生态环境保护办公室": ["房屋建筑", "交通工程", "人防工程", "园林绿化工程", "既有建筑修缮", "既有建筑加装电梯工程", "临时建设", "农民自建房", "建筑装饰装修工程"],
    "朱家角镇-城市精细化管理工作部": ["住宅室内装饰装修工程"],
    "朱家角镇-经济发展办公室（农业农村发展办公室）": ["建筑装饰装修工程"],
    "朱家角镇-土地管理工作部": ["农民自建房"],
    "朱家角镇-农业农村服务中心": ["农业设施建设工程", "农村地区综合性建设工程"],

    # 白鹤镇
    "白鹤镇-城市建设管理事务中心": ["住宅室内装饰装修工程", "既有建筑修缮", "既有建筑加装电梯工程", "水务工程"],
    "白鹤镇-规划建设和生态环境保护办公室": ["房屋建筑", "交通工程", "人防工程", "农民自建房", "建筑装饰装修工程", "既有建筑修缮", "临时建设"],
    "白鹤镇-农业农村服务中心": ["农业设施建设工程", "农村地区综合性建设工程", "园林绿化工程"],
    "白鹤镇-社会事业发展办公室": ["园林绿化工程"],

    # 盈浦
    "盈浦街道-城市建设管理事务中心": ["临时建设", "住宅室内装饰装修工程", "农业设施建设工程", "农村地区综合性建设工程", "农民自建房", "建筑装饰装修工程", "房屋建筑", "既有建筑修缮", "既有建筑加装电梯工程", "水务工程"],
    "盈浦街道-社区管理办公室": ["交通工程", "人防工程", "园林绿化工程"],
    "盈浦街道-城市运行管理中心": ["交通工程"],
    "盈浦街道-武装部": ["人防工程"],
    "盈浦街道-经济服务办公室": ["既有建筑修缮"],
    "区市场监管局-盈浦市场监督管理所": ["既有建筑修缮"],

    # 练塘镇
    "练塘镇-城市建设管理事务中心": ["住宅室内装饰装修工程", "水务工程", "既有建筑修缮", "既有建筑加装电梯工程"],
    "练塘镇-规划建设和生态环境保护办公室": ["房屋建筑", "交通工程", "建筑装饰装修工程", "既有建筑修缮", "既有建筑加装电梯工程", "临时建设"],
    "练塘镇-社会事业发展办公室": ["园林绿化工程"],
    "练塘镇-武装部": ["人防工程"],
    "练塘镇-城乡建设工作部": ["农民自建房"],
    "练塘镇-农业农村服务中心": ["农业设施建设工程"],
    "练塘镇-新农村建设工作专班": ["农村地区综合性建设工程"],

    # 赵巷镇
    "赵巷镇-城市建设管理事务中心": ["临时建设", "住宅室内装饰装修工程", "既有建筑修缮", "既有建筑加装电梯工程", "水务工程"],
    "赵巷镇-规划建设和生态环境办公室": ["交通工程", "农民自建房", "建筑装饰装修工程", "房屋建筑", "既有建筑修缮", "临时建设"],
    "赵巷镇-社会事业发展办公室": ["园林绿化工程"],
    "赵巷镇-武装部": ["人防工程"],
    "赵巷镇-农业农村服务中心": ["农业设施建设工程", "农村地区综合性建设工程"],

    # 重固镇
    "重固镇-城市建设管理事务中心": ["住宅室内装饰装修工程", "既有建筑修缮", "水务工程"],
    "重固镇-规划建设和生态环境保护办公室": ["房屋建筑", "交通工程", "农民自建房", "建筑装饰装修工程", "既有建筑修缮", "临时建设", "水务工程"],
    "重固镇-社会事业发展办公室": ["园林绿化工程"],
    "重固镇-武装部": ["人防工程"],
    "重固镇-农业农村服务中心": ["农业设施建设工程", "农村地区综合性建设工程"],
    "区市场监管局-重固市场监督管理所": ["既有建筑加装电梯工程"],

    # 金泽镇
    "金泽镇-城市建设管理事务中心": ["住宅室内装饰装修工程", "既有建筑修缮", "既有建筑加装电梯工程", "水务工程"],
    "金泽镇-规划建设和生态环境保护办公室": ["临时建设", "交通工程", "农民自建房", "房屋建筑", "既有建筑修缮"],
    "金泽镇-社会事业发展办公室": ["园林绿化工程"],
    "金泽镇-武装部": ["人防工程"],
    "金泽镇-农业农村服务中心": ["农业设施建设工程"],
    "金泽镇-乡村振兴办": ["农村地区综合性建设工程"],
    "金泽镇-经济发展办公室": ["建筑装饰装修工程", "既有建筑加装电梯工程"],

    # 青浦工业园区
    "青浦工业园区-规划建设部": ["临时建设", "建筑装饰装修工程", "房屋建筑", "既有建筑修缮", "既有建筑加装电梯工程"],

    # 香花桥
    "香花桥-城市建设管理事务中心": ["人防工程", "住宅室内装饰装修工程", "农业设施建设工程", "农村地区综合性建设工程", "既有建筑修缮", "既有建筑加装电梯工程", "水务工程"],
    "香花桥-社区管理办公室": ["临时建设", "交通工程", "农民自建房", "园林绿化工程", "建筑装饰装修工程", "房屋建筑", "既有建筑修缮"],
}


def parse_types(type_str):
    """解析工程类型字符串为集合"""
    if not type_str:
        return set()
    return set(t.strip() for t in str(type_str).split("，") if t.strip())


def format_types(type_set):
    """将类型集合格式化为"，"分隔的字符串"""
    # 保持特定顺序
    order = ["临时建设", "交通工程", "人防工程", "园林绿化工程", "农业设施建设工程",
             "农村地区综合性建设工程", "农民自建房", "建筑装饰装修工程", "房屋建筑",
             "既有建筑修缮", "既有建筑加装电梯工程", "住宅室内装饰装修工程", "水务工程"]
    result = []
    for t in order:
        if t in type_set:
            result.append(t)
    # 添加不在order中的其他类型
    for t in sorted(type_set):
        if t not in order:
            result.append(t)
    return "，".join(result) if result else ""


def audit_file(filepath):
    """核对单个文件"""
    print(f"\n{'='*70}")
    print(f"核对文件: {os.path.basename(filepath)}")
    print(f"{'='*70}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb["组织导入"]

    # 找列
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[h] = col

    org_col = headers.get("组织名称")
    type_col = headers.get("可审核工程类型")

    issues = []
    auto_fixes = []

    for row in range(2, ws.max_row + 1):
        org = ws.cell(row=row, column=org_col).value
        if not org or org not in EXPECTED_TYPES:
            continue

        actual = parse_types(ws.cell(row=row, column=type_col).value)
        expected = set(EXPECTED_TYPES[org])

        missing = expected - actual  # 期望有但实际没有的
        extra = actual - expected     # 实际有但期望没有的

        if missing or extra:
            issues.append({
                "row": row,
                "org": org,
                "missing": missing,
                "extra": extra,
                "actual": actual,
                "expected": expected,
            })

    if not issues:
        print("✓ 没有发现差异，所有组织的可审核工程类型与底稿一致！")
        return []

    print(f"\n发现 {len(issues)} 个组织有差异：\n")

    for issue in issues:
        org = issue["org"]
        print(f"  【{org}】")
        if issue["missing"]:
            print(f"    遗漏: {', '.join(sorted(issue['missing']))}")
        if issue["extra"]:
            print(f"    多余: {', '.join(sorted(issue['extra']))}")
        print(f"    当前: {format_types(issue['actual'])}")
        print(f"    应为: {format_types(issue['expected'])}")
        print()
        auto_fixes.append(issue)

    return auto_fixes


def fix_file(filepath, issues):
    """自动修正文件"""
    print(f"\n开始自动修正: {os.path.basename(filepath)}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb["组织导入"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[h] = col

    org_col = headers.get("组织名称")
    type_col = headers.get("可审核工程类型")

    for issue in issues:
        row = issue["row"]
        new_types = format_types(issue["expected"])
        ws.cell(row=row, column=type_col).value = new_types
        print(f"  修正: {issue['org']} → {new_types}")

    wb.save(filepath)
    print(f"  已保存！")


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))

    for fname in ["导入模板_最终版.xlsx", "导入模板_已处理.xlsx"]:
        fpath = os.path.join(base_dir, fname)
        if os.path.exists(fpath):
            issues = audit_file(fpath)
            if issues:
                fix_file(fpath, issues)
        else:
            print(f"文件不存在: {fpath}")

    print("\n" + "="*70)
    print("核对完成!")
    print("="*70)
