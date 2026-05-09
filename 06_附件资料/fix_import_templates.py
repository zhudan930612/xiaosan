#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
全面修正导入模板：
1. 同步组织名称（简称→全称）
2. 去掉"住宅非住宅备注"列
3. 修正小型工程作业场所（非住宅T3-5/T3-7 → "除住宅小区公共区域内场所以外的其他场所"）
4. 修正可审核工程类型（遗漏的添加，错误的移除）
5. 同步修正权限判断明细表中的组织条款
"""

import openpyxl
import os
import copy

# ========== 组织名称映射（旧简称 → 新全称） ==========
# 用于人员导入表和权限判断明细表中的组织名称同步
ORG_NAME_MAP = {
    # 西虹桥
    "上海西虹桥商务开发有限公司-建设管理部": "西虹桥公司-建设管理部",
    # 华新镇
    "华新镇-城建中心": "华新镇-城市建设管理事务中心",
    "华新镇-市场监督管理所": "区市场监管局-华新市场监督管理所",
    "华新镇-绿化公司": "华新镇-镇属企业-绿化公司",
    "华新镇-规建办": "华新镇-规划建设和生态环境办公室",
    # 朱家角镇
    "朱家角镇-土地部": "朱家角镇-土地管理工作部",
    "朱家角镇-城建中心": "朱家角镇-城市建设管理事务中心",
    "朱家角镇-精细化部": "朱家角镇-城市精细化管理工作部",
    "朱家角镇-经发办": "朱家角镇-经济发展办公室（农业农村发展办公室）",
    "朱家角镇-规建办": "朱家角镇-规划建设和生态环境保护办公室",
    "朱家角镇-农服中心": "朱家角镇-农业农村服务中心",
    # 白鹤镇
    "白鹤镇-城建中心": "白鹤镇-城市建设管理事务中心",
    "白鹤镇-社发办": "白鹤镇-社会事业发展办公室",
    "白鹤镇-规保办": "白鹤镇-规划建设和生态环境保护办公室",
    "白鹤镇-农服中心": "白鹤镇-农业农村服务中心",
    # 盈浦
    "盈浦-城建中心": "盈浦街道-城市建设管理事务中心",
    "盈浦-城运中心": "盈浦街道-城市运行管理中心",
    "盈浦-市场所": "区市场监管局-盈浦市场监督管理所",
    "盈浦-武装部": "盈浦街道-武装部",
    "盈浦-管理办": "盈浦街道-社区管理办公室",
    "盈浦-经服办": "盈浦街道-经济服务办公室",
    # 练塘镇
    "练塘镇-城乡建设部": "练塘镇-城乡建设工作部",
    "练塘镇-城建中心": "练塘镇-城市建设管理事务中心",
    "练塘镇-社发办": "练塘镇-社会事业发展办公室",
    "练塘镇-规建生态办": "练塘镇-规划建设和生态环境保护办公室",
    "练塘镇-农服中心": "练塘镇-农业农村服务中心",
    # 赵巷镇
    "赵巷镇-城建中心": "赵巷镇-城市建设管理事务中心",
    "赵巷镇-社发办": "赵巷镇-社会事业发展办公室",
    "赵巷镇-规建办": "赵巷镇-规划建设和生态环境办公室",
    "赵巷镇-农服中心": "赵巷镇-农业农村服务中心",
    # 重固镇
    "重固镇-城建中心": "重固镇-城市建设管理事务中心",
    "重固镇-市场所": "区市场监管局-重固市场监督管理所",
    "重固镇-社发办": "重固镇-社会事业发展办公室",
    "重固镇-规生办": "重固镇-规划建设和生态环境保护办公室",
    "重固镇-农服中心": "重固镇-农业农村服务中心",
    # 金泽镇
    "金泽镇-城建中心": "金泽镇-城市建设管理事务中心",
    "金泽镇-社发办": "金泽镇-社会事业发展办公室",
    "金泽镇-经发办": "金泽镇-经济发展办公室",
    "金泽镇-规建办": "金泽镇-规划建设和生态环境保护办公室",
    "金泽镇-农服中心": "金泽镇-农业农村服务中心",
    # 香花桥
    "香花桥-城建中心": "香花桥-城市建设管理事务中心",
    "香花桥-社区管理办": "香花桥-社区管理办公室",
}

NON_RESIDENTIAL_PLACE = "除住宅小区公共区域内场所以外的其他场所"

# 只负责非住宅T3-5/T3-7的组织（新全称）→ 场所应改为非住宅
NON_RESIDENTIAL_ORGS = [
    "华新镇-规划建设和生态环境办公室",      # T3-5非住宅
    "区市场监管局-华新市场监督管理所",      # T3-7非住宅
    "白鹤镇-规划建设和生态环境保护办公室",  # T3-5非住宅
    "练塘镇-规划建设和生态环境保护办公室",  # T3-5非住宅, T3-7非住宅
    "赵巷镇-规划建设和生态环境办公室",      # T3-5非住宅
    "金泽镇-规划建设和生态环境保护办公室",  # T3-5非住宅
    "金泽镇-经济发展办公室",                  # T3-7非住宅
    "香花桥-社区管理办公室",                  # T3-5非住宅
    "区市场监管局-盈浦市场监督管理所",      # T3-5非住宅
]


def replace_org_name(text):
    """替换文本中的组织名称（从旧简称到新全称）"""
    if not text:
        return text
    for old_name in sorted(ORG_NAME_MAP.keys(), key=len, reverse=True):
        new_name = ORG_NAME_MAP[old_name]
        text = text.replace(old_name, new_name)
    return text


def fix_workbook(filepath, is_final_version=True):
    """修正单个Excel工作簿"""
    print(f"\n{'='*60}")
    print(f"处理文件: {os.path.basename(filepath)}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(filepath)
    changes = []

    # ========== 1. 组织导入表 ==========
    if "组织导入" in wb.sheetnames:
        ws = wb["组织导入"]
        print("\n【组织导入表】")

        # 找列
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[h] = col

        org_col = headers.get("组织名称")
        type_col = headers.get("可审核工程类型")
        place_col = headers.get("小型工程作业场所")
        remark_col = headers.get("住宅非住宅备注")

        # 注：用户要求暂时保留"住宅非住宅备注"列
        # if remark_col:
        #     ws.delete_cols(remark_col)
        #     changes.append(f"删除'住宅非住宅备注'列")

        # 逐行修正
        for row in range(2, ws.max_row + 1):
            org_name = ws.cell(row=row, column=org_col).value
            if not org_name:
                continue

            # 1) 修正场所：非住宅组织
            if org_name in NON_RESIDENTIAL_ORGS and place_col:
                old_place = ws.cell(row=row, column=place_col).value
                if old_place != NON_RESIDENTIAL_PLACE:
                    ws.cell(row=row, column=place_col).value = NON_RESIDENTIAL_PLACE
                    changes.append(f"组织导入 行{row}: {org_name} 场所 '{old_place}' → '{NON_RESIDENTIAL_PLACE}'")

            # 2) 修正可审核工程类型
            if type_col:
                old_types = ws.cell(row=row, column=type_col).value or ""
                new_types = old_types

                # 盈浦经服办：添加既有建筑修缮
                if org_name == "盈浦街道-经济服务办公室":
                    if "既有建筑修缮" not in new_types:
                        new_types = (new_types + "，既有建筑修缮").strip("，")
                        changes.append(f"组织导入 行{row}: {org_name} 添加'既有建筑修缮'")

                # 盈浦城运中心：移除既有建筑修缮
                elif org_name == "盈浦街道-城市运行管理中心":
                    if "既有建筑修缮" in new_types:
                        parts = [p.strip() for p in new_types.split("，") if p.strip() != "既有建筑修缮"]
                        new_types = "，".join(parts)
                        changes.append(f"组织导入 行{row}: {org_name} 移除'既有建筑修缮'")

                # 练塘城建中心：添加既有建筑修缮和既有建筑加装电梯工程
                elif org_name == "练塘镇-城市建设管理事务中心":
                    added = []
                    if "既有建筑修缮" not in new_types:
                        new_types = (new_types + "，既有建筑修缮").strip("，")
                        added.append("既有建筑修缮")
                    if "既有建筑加装电梯工程" not in new_types:
                        new_types = (new_types + "，既有建筑加装电梯工程").strip("，")
                        added.append("既有建筑加装电梯工程")
                    if added:
                        changes.append(f"组织导入 行{row}: {org_name} 添加{added}")

                if new_types != old_types:
                    ws.cell(row=row, column=type_col).value = new_types

    # ========== 2. 人员导入表 ==========
    if "人员导入" in wb.sheetnames:
        ws = wb["人员导入"]
        print("\n【人员导入表】")

        org_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value in ("所属组织名称", "组织名称"):
                org_col = col
                break

        if org_col:
            count = 0
            for row in range(2, ws.max_row + 1):
                org = ws.cell(row=row, column=org_col).value
                if org:
                    new_org = replace_org_name(org)
                    if new_org != org:
                        ws.cell(row=row, column=org_col).value = new_org
                        count += 1
            if count > 0:
                changes.append(f"人员导入表: 同步{count}行组织名称")

    # ========== 3. 权限判断明细表 ==========
    if "权限判断明细" in wb.sheetnames:
        ws = wb["权限判断明细"]
        print("\n【权限判断明细表】")

        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[h] = col

        org_col = headers.get("所属组织(解析)")
        raw_org_col = headers.get("原始组织")
        remark_col = headers.get("住宅非住宅备注")
        org_type_col = headers.get("组织条款")

        # 注：用户要求暂时保留"住宅非住宅备注"列
        # if remark_col:
        #     ws.delete_cols(remark_col)
        #     changes.append(f"权限判断明细: 删除'住宅非住宅备注'列")

        # 同步组织名称 + 修正组织条款
        name_count = 0
        for row in range(2, ws.max_row + 1):
            # 同步组织名称
            if org_col:
                org = ws.cell(row=row, column=org_col).value
                if org:
                    new_org = replace_org_name(org)
                    if new_org != org:
                        ws.cell(row=row, column=org_col).value = new_org
                        name_count += 1

            if raw_org_col:
                raw = ws.cell(row=row, column=raw_org_col).value
                if raw:
                    new_raw = replace_org_name(raw)
                    if new_raw != raw:
                        ws.cell(row=row, column=raw_org_col).value = new_raw

            # 修正组织条款
            if org_type_col and org_col:
                org = ws.cell(row=row, column=org_col).value
                old_clause = ws.cell(row=row, column=org_type_col).value or ""
                new_clause = old_clause

                if org == "练塘镇-城市建设管理事务中心":
                    if "既有建筑修缮" not in new_clause:
                        new_clause = (new_clause + "，既有建筑修缮").strip("，")
                    if "既有建筑加装电梯工程" not in new_clause:
                        new_clause = (new_clause + "，既有建筑加装电梯工程").strip("，")
                    if new_clause != old_clause:
                        ws.cell(row=row, column=org_type_col).value = new_clause
                        changes.append(f"权限判断 行{row}: {org} 组织条款添加'既有建筑修缮、既有建筑加装电梯工程'")

                elif org == "盈浦街道-城市运行管理中心":
                    if "既有建筑修缮" in new_clause:
                        parts = [p.strip() for p in new_clause.split("，") if p.strip() != "既有建筑修缮"]
                        new_clause = "，".join(parts)
                        ws.cell(row=row, column=org_type_col).value = new_clause
                        changes.append(f"权限判断 行{row}: {org} 组织条款移除'既有建筑修缮'")

        if name_count > 0:
            changes.append(f"权限判断明细表: 同步{name_count}行组织名称")

    # 保存
    wb.save(filepath)
    print(f"\n修改完成，共{len(changes)}处变更：")
    for c in changes:
        print(f"  - {c}")
    return changes


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))

    for fname in ["导入模板_最终版.xlsx", "导入模板_已处理.xlsx"]:
        fpath = os.path.join(base_dir, fname)
        if os.path.exists(fpath):
            fix_workbook(fpath)
        else:
            print(f"文件不存在: {fpath}")

    print("\n" + "="*60)
    print("全部处理完成!")
    print("="*60)
