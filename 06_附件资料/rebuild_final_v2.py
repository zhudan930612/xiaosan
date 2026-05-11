#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从导入模板_最终版.xlsx 重新生成 导入模板_最终版2.xlsx
确保人员导入表数据正确
"""

import openpyxl
import shutil
import os

src = '导入模板_最终版.xlsx'
dst = '导入模板_最终版2.xlsx'

# 复制文件
if os.path.exists(dst):
    os.remove(dst)
shutil.copy2(src, dst)
print(f'已复制: {src} → {dst}')

src_wb = openpyxl.load_workbook(src)
dst_wb = openpyxl.load_workbook(dst)

# ===== 1. 组织导入表 =====
ws = dst_wb['组织导入']
headers = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

org_col = headers['组织名称']
area_col = headers['管辖区域']
type_col = headers['可审核工程类型']
small_place_col = headers['小型工程作业场所']
scattered_place_col = headers['零星作业作业场所']

ALL_SMALL = {
    '新改扩房屋建设工程', '交通工程', '水务工程', '园林绿化工程',
    '既有建筑修缮工程', '建筑装饰装修工程', '既有建筑加装电梯工程',
    '临时建设工程', '农民自建房工程', '农业设施建设工程',
    '农村地区综合性建设工程', '人防工程',
}

# 读取所有数据
rows_data = []
for row in range(2, ws.max_row + 1):
    org = ws.cell(row=row, column=org_col).value
    area = ws.cell(row=row, column=area_col).value
    types = ws.cell(row=row, column=type_col).value or ''
    small_place = ws.cell(row=row, column=small_place_col).value
    scattered_place = ws.cell(row=row, column=scattered_place_col).value
    rows_data.append((org, area, types, small_place, scattered_place))

# 清空工作表并重建
ws.delete_rows(1, ws.max_row)
ws.append(['组织名称', '管辖区域', '小型工程类型', '零星作业类型', '住宅装修类型', '小型工程作业场所', '零星作业作业场所'])

for org, area, types, small_place, scattered_place in rows_data:
    parts = set(t.strip() for t in types.split('，') if t.strip())
    deco = '住宅室内装饰装修工程' if '住宅室内装饰装修工程' in parts else ''
    small = parts - {'住宅室内装饰装修工程'}

    if small == ALL_SMALL:
        small_str = '全部'
    else:
        order = [
            '新改扩房屋建设工程', '交通工程', '水务工程', '园林绿化工程',
            '既有建筑修缮工程', '建筑装饰装修工程', '既有建筑加装电梯工程',
            '临时建设工程', '农民自建房工程', '农业设施建设工程',
            '农村地区综合性建设工程', '人防工程',
        ]
        small_str = '，'.join(t for t in order if t in small)

    ws.append([org, area, small_str, '', deco, small_place, scattered_place])

print('组织导入表已重建')

# ===== 2. 人员导入表 =====
ws2 = dst_wb['人员导入']
headers2 = {ws2.cell(row=1, column=col).value: col for col in range(1, ws2.max_column + 1)}

name_col = headers2['姓名']
phone_col = headers2['手机号']
org_name_col = headers2['所属组织名称']
role_col = headers2['角色']
perm_col = headers2['权限模式']
type_col_p = headers2['自定义工程类型']
place_col_p = headers2['自定义作业场所']

# 读取所有数据
person_data = []
for row in range(2, ws2.max_row + 1):
    person_data.append({
        'name': ws2.cell(row=row, column=name_col).value,
        'phone': ws2.cell(row=row, column=phone_col).value,
        'org': ws2.cell(row=row, column=org_name_col).value,
        'role': ws2.cell(row=row, column=role_col).value,
        'perm': ws2.cell(row=row, column=perm_col).value,
        'types': ws2.cell(row=row, column=type_col_p).value or '',
        'place': ws2.cell(row=row, column=place_col_p).value or '',
    })

# 清空并重建
ws2.delete_rows(1, ws2.max_row)
ws2.append(['姓名', '手机号', '所属组织名称', '角色', '权限模式',
            '自定义小型工程类型', '自定义住宅装修类型',
            '自定义小型工程作业场所', '自定义零星作业场所'])

for p in person_data:
    parts = [x.strip() for x in p['types'].split('，') if x.strip()]
    small_types = []
    deco_types = []
    for part in parts:
        if part == '住宅室内装饰装修工程':
            deco_types.append(part)
        else:
            small_types.append(part)

    ws2.append([
        p['name'], p['phone'], p['org'], p['role'], p['perm'],
        '，'.join(small_types) if small_types else '',
        '，'.join(deco_types) if deco_types else '',
        p['place'],
        '',
    ])

print('人员导入表已重建')

# 保存
dst_wb.save(dst)
print(f'已保存: {dst}')

# 验证
print('\n组织导入表列:')
for col in range(1, dst_wb['组织导入'].max_column + 1):
    print(f'  {col}: {dst_wb["组织导入"].cell(row=1, column=col).value}')

print('\n人员导入表列:')
for col in range(1, dst_wb['人员导入'].max_column + 1):
    print(f'  {col}: {dst_wb["人员导入"].cell(row=1, column=col).value}')
