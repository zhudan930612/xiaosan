#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从"导入模板_最终版.xlsx"生成"导入模板_最终版2.xlsx"
重新设计组织导入表列结构：
  组织名称 | 管辖区域 | 小型工程类型 | 零星作业类型 | 住宅装修类型 | 小型工程作业场所 | 零星作业作业场所
"""

import openpyxl
import shutil
import os

# 12种小型工程类型（标准顺序）
ALL_SMALL_TYPES = {
    "新改扩房屋建设工程",
    "交通工程",
    "水务工程",
    "园林绿化工程",
    "既有建筑修缮工程",
    "建筑装饰装修工程",
    "既有建筑加装电梯工程",
    "临时建设工程",
    "农民自建房工程",
    "农业设施建设工程",
    "农村地区综合性建设工程",
    "人防工程",
}


def create_v2():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    src = os.path.join(base_dir, "导入模板_最终版.xlsx")
    dst = os.path.join(base_dir, "导入模板_最终版2.xlsx")

    # 复制文件
    shutil.copy2(src, dst)
    print(f"已复制: {src} → {dst}")

    wb = openpyxl.load_workbook(dst)
    ws = wb["组织导入"]

    # 找原列
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[h] = col

    org_col = headers.get("组织名称")
    area_col = headers.get("管辖区域")
    type_col = headers.get("可审核工程类型")
    small_place_col = headers.get("小型工程作业场所")
    scattered_place_col = headers.get("零星作业作业场所")

    # 删除原"可审核工程类型"列，在其位置插入3列新列
    ws.insert_cols(type_col, 3)

    # 新的列结构（type_col现在是"可审核工程类型"列的位置，但插入3列后，type_col及之后的列号都变了）
    # type_col 位置现在是新的C列位置
    # 我们需要在type_col、type_col+1、type_col+2位置写入新列名
    ws.cell(row=1, column=type_col).value = "小型工程类型"
    ws.cell(row=1, column=type_col + 1).value = "零星作业类型"
    ws.cell(row=1, column=type_col + 2).value = "住宅装修类型"

    # 新的small_place_col和scattered_place_col列号也变了（往后移了3列）
    new_small_place_col = small_place_col + 3
    new_scattered_place_col = scattered_place_col + 3

    # 处理每一行
    for row in range(2, ws.max_row + 1):
        # 从原始可审核工程类型（现在在type_col+3列）拆分
        original_types_str = ws.cell(row=row, column=type_col + 3).value or ""
        original_types = set(t.strip() for t in original_types_str.split("，") if t.strip())

        # 分离住宅装修类型
        decoration_type = "住宅室内装饰装修工程" if "住宅室内装饰装修工程" in original_types else ""

        # 分离小型工程类型
        small_types = original_types - {"住宅室内装饰装修工程"}

        # 判断是否全部12种
        if small_types == ALL_SMALL_TYPES:
            small_type_str = "全部"
        else:
            # 按标准顺序排列
            order = [
                "新改扩房屋建设工程", "交通工程", "水务工程", "园林绿化工程",
                "既有建筑修缮工程", "建筑装饰装修工程", "既有建筑加装电梯工程",
                "临时建设工程", "农民自建房工程", "农业设施建设工程",
                "农村地区综合性建设工程", "人防工程",
            ]
            parts = [t for t in order if t in small_types]
            small_type_str = "，".join(parts)

        # 写入新列
        ws.cell(row=row, column=type_col).value = small_type_str
        ws.cell(row=row, column=type_col + 1).value = ""  # 零星作业类型全部留空
        ws.cell(row=row, column=type_col + 2).value = decoration_type

    # 删除原"可审核工程类型"列（现在在type_col+3列）
    ws.delete_cols(type_col + 3)

    # 保存
    wb.save(dst)
    print(f"已保存: {dst}")
    print("\n新列结构:")
    for col in range(1, ws.max_column + 1):
        print(f"  {col}: {ws.cell(row=1, column=col).value}")


if __name__ == "__main__":
    create_v2()
    print("\n完成!")
