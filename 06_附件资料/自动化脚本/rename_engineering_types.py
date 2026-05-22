#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将导入模板中的工程类型名称改为完整规范名称：
- 房屋建筑 → 新改扩房屋建设工程
- 既有建筑修缮 → 既有建筑修缮工程
- 临时建设 → 临时建设工程
- 农民自建房 → 农民自建房工程
"""

import openpyxl
import os

# 工程类型重命名（精确匹配，避免重复替换）
TYPE_RENAME = {
    "房屋建筑": "新改扩房屋建设工程",
    "既有建筑修缮": "既有建筑修缮工程",
    "临时建设": "临时建设工程",
    "农民自建房": "农民自建房工程",
}


def rename_types_in_str(s):
    """对'，'分隔的工程类型字符串做精确重命名（按拆分后的元素匹配，避免误改）"""
    if not s or not isinstance(s, str):
        return s
    # 兼容中英文逗号
    parts = []
    for p in s.replace(",", "，").split("，"):
        p = p.strip()
        if not p:
            continue
        parts.append(TYPE_RENAME.get(p, p))
    return "，".join(parts)


def fix_workbook(filepath):
    """处理单个工作簿"""
    print(f"\n{'='*60}")
    print(f"处理文件: {os.path.basename(filepath)}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(filepath)
    total_changes = 0

    # ===== 组织导入：可审核工程类型 =====
    if "组织导入" in wb.sheetnames:
        ws = wb["组织导入"]
        type_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "可审核工程类型":
                type_col = col
                break
        if type_col:
            for row in range(2, ws.max_row + 1):
                old = ws.cell(row=row, column=type_col).value
                if old:
                    new = rename_types_in_str(old)
                    if new != old:
                        ws.cell(row=row, column=type_col).value = new
                        total_changes += 1

    # ===== 人员导入：自定义工程类型 =====
    if "人员导入" in wb.sheetnames:
        ws = wb["人员导入"]
        type_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "自定义工程类型":
                type_col = col
                break
        if type_col:
            for row in range(2, ws.max_row + 1):
                old = ws.cell(row=row, column=type_col).value
                if old:
                    new = rename_types_in_str(old)
                    if new != old:
                        ws.cell(row=row, column=type_col).value = new
                        total_changes += 1

    # ===== 权限判断明细：组织条款、人员条款、自定义工程类型 =====
    if "权限判断明细" in wb.sheetnames:
        ws = wb["权限判断明细"]
        target_cols = []
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col).value
            if h in ("组织条款", "人员条款", "自定义工程类型"):
                target_cols.append(col)
        for col in target_cols:
            for row in range(2, ws.max_row + 1):
                old = ws.cell(row=row, column=col).value
                if old:
                    new = rename_types_in_str(old)
                    if new != old:
                        ws.cell(row=row, column=col).value = new
                        total_changes += 1

    wb.save(filepath)
    print(f"完成，共修改 {total_changes} 处")
    return total_changes


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    grand_total = 0
    for fname in ["导入模板_最终版.xlsx", "导入模板_已处理.xlsx"]:
        fpath = os.path.join(base_dir, fname)
        if os.path.exists(fpath):
            grand_total += fix_workbook(fpath)
        else:
            print(f"文件不存在: {fpath}")
    print(f"\n{'='*60}")
    print(f"全部完成! 总计修改 {grand_total} 处")
    print(f"{'='*60}")
